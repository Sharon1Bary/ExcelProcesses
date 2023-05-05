from src.lib.domains.Core import Core

import openpyxl


def find_regions_utils(search_term: str, core: Core) -> list:

    files = []
    regions = []
    for current_category in core.categories.values():
        files.append({'region': current_category.region, 'files': current_category.files})

    for current in files:
        for file in current['files']:
            i = 1
            ws = openpyxl.load_workbook(file).active

            while i <= ws.max_column:
                cells = next(ws.iter_cols(min_col=i, values_only=True))
                if search_term in cells:
                    regions.append(current['region'])
                i += 1
    return regions
