from src.lib.domains.Core import Core
import openpyxl


def sum_type_utils(type_: str, core: Core) -> int:

    files = []
    sum_ = 0
    for category in core.categories.values():
        if category.type_ == type_:
            files.append(category.files)
    for current_files in files:
        sum_ += sum_from_file(current_files, sum_)
    return sum_


def sum_from_file(files: [], sum_: int) -> int:

    for file in files:
        i = 1
        all_cells = []
        ws = openpyxl.load_workbook(file).active
        while i <= ws.max_column:
            all_cells.append(next(ws.iter_cols(min_col=i, values_only=True)))
            i += 1
        sum_ += sum_file_cells(cells=all_cells)
    return sum_


def sum_file_cells(cells: list) -> int:
    return sum([sum(x) for x in [tuple(filter(lambda x:  isinstance(x, int), lis)) for lis in cells]])

    # res = [tuple(filter(lambda x:  isinstance(x, int), lis)) for lis in cells] returns list of tuples. [(),()]
    # [sum(x) for x in [tuple(filter(lambda x: res] returns list on sum in each tuple.
